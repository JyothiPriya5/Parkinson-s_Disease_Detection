from flask import Flask, render_template, request, redirect, url_for, flash
import numpy as np
import pickle
import openpyxl
import os
import joblib
from werkzeug.security import generate_password_hash, check_password_hash

# Initialization of Flask object
app = Flask(__name__)
app.secret_key = 'your_secret_key'  # Set a secret key for flash messages

# Load ML model
try:
    model = pickle.load(open('rf_clf.pkl', 'rb'))
except FileNotFoundError:
    print("Model file 'rf_clf.pkl' not found. Ensure the file is in the directory.")

try:
    scaler = joblib.load('minmax_scaler.pkl')
except FileNotFoundError:
    print("MinMax scaler file 'minmax_scaler.pkl' not found. Ensure the file is in the directory.")


# Define the Excel file paths
user_data_file = 'user_data.xlsx'
prediction_data_file = 'prediction_data.xlsx'

# Initialize the Excel files with headers if they don't exist
def initialize_excel_file(file_path, headers):
    if not os.path.exists(file_path):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(headers)
        workbook.save(file_path)

initialize_excel_file(user_data_file, ['Email', 'Password'])
initialize_excel_file(prediction_data_file, ['Specific Gravity', 'Hypertension', 'Hemoglobin', 'Diabetes Mellitus', 'Albumin', 'Appetite', 'Red Blood Cells', 'Pus Cell'])


@app.route('/', methods=['GET', 'POST'])
def home():
    if request.method == 'POST':
        email = request.form['loginEmail']
        password = request.form['loginPassword']
        
        # Check user credentials
        workbook = openpyxl.load_workbook(user_data_file)
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == email and check_password_hash(row[1], password):
                flash('Login successful!', 'success')
                return redirect(url_for('index'))
        
        flash('Invalid email or password', 'error')
    
    return render_template('login_signup.html')


@app.route('/signup', methods=['POST'])
def signup():
    name = request.form['signupName']
    email = request.form['signupEmail']
    password = request.form['signupPassword']
    confirm_password = request.form['signupConfirmPassword']

    if password != confirm_password:
        flash('Passwords do not match', 'error')
        return redirect(url_for('home'))

    # Check if email already exists
    workbook = openpyxl.load_workbook(user_data_file)
    sheet = workbook.active
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] == email:
            flash('Email already in use', 'error')
            return redirect(url_for('home'))

    # Add new user with hashed password
    hashed_password = generate_password_hash(password)
    sheet.append([email, hashed_password])
    workbook.save(user_data_file)

    flash('Signup successful! You can now log in.', 'success')
    return redirect(url_for('home'))


@app.route('/index')
def index():
    return render_template('index.html')


@app.route('/admin', methods=['GET', 'POST'])
def admin():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        if username == 'admin' and password == 'admin':
            # Load the Excel file and retrieve user data
            workbook = openpyxl.load_workbook(prediction_data_file)
            sheet = workbook.active
            data = [row for row in sheet.iter_rows(min_row=2, values_only=True)]
            return render_template('admin_data.html', data=data)
        else:
            flash('Invalid admin credentials', 'error')
    
    return render_template('admin.html')


@app.route("/predict", methods=['POST'])
def predict():
    if request.method == 'POST':
        # Retrieve inputs and convert to float
        try:
            inputs = [
                float(request.form['MDVP:Fo(Hz)']),
                float(request.form['MDVP:Fhi(Hz)']),
                float(request.form['MDVP:Flo(Hz)']),
                float(request.form['MDVP:Jitter(%)']),
                float(request.form['MDVP:Jitter(Abs)']),
                float(request.form['MDVP:RAP']),
                float(request.form['MDVP:PPQ']),
                float(request.form['Jitter:DDP']),
                float(request.form['MDVP:Shimmer']),
                float(request.form['MDVP:Shimmer(dB)']),
                float(request.form['Shimmer:APQ3']),
                float(request.form['Shimmer:APQ5']),
                float(request.form['MDVP:APQ']),
                float(request.form['Shimmer:DDA']),
                float(request.form['NHR']),
                float(request.form['HNR']),
                float(request.form['RPDE']),
                float(request.form['DFA']),
                float(request.form['spread1']),
                float(request.form['spread2']),
                float(request.form['D2']),
                float(request.form['PPE'])
            ]
        except ValueError:
            flash("Invalid input. Please enter numeric values.", 'error')
            return redirect(url_for('index'))
        
        # Save the user input to the Excel file
        workbook = openpyxl.load_workbook(prediction_data_file)
        sheet = workbook.active
        sheet.append(inputs)
        workbook.save(prediction_data_file)

        # Predict using the model
        values = np.array([inputs])
        scaled_values = scaler.transform(values)
        prediction = model.predict(scaled_values)
        print(prediction)
        print("hi")

        return render_template('result.html', prediction=prediction)


if __name__ == "__main__":
    app.run(debug=True)
